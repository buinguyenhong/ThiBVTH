import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CATALOG_DIR = os.path.join(BASE_DIR, "data", "catalogs")

def create_folders():
    os.makedirs(CATALOG_DIR, exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, "data", "output"), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, "data", "config"), exist_ok=True)

def create_patients_catalog():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    ws.append(["MaYTe", "TenBenhNhan", "NgaySinh", "GioiTinh", "DiaChiLienHe", "SoBHYT", "BHYTTuNgay", "BHYTDenNgay", "DKKCB"])
    patients = [
        ["26034346", "Lê Rất Vui", "02/11/1991", "Nữ", "Thôn 1, Xã Hòa Phú, TP. Buôn Ma Thuột, Đắk Lắk", "DN4666622153422", "01/01/2024", "31/12/2026", "66232"],
        ["26000270", "Nguyễn Thị Hường", "07/02/1996", "Nữ", "Thôn 10, Xã Nam Bình, Huyện Đắk Song, Đắk Nông", "DN4666624113158", "01/01/2024", "31/12/2026", "66232"],
        ["1127779", "Đồng Viết Thanh", "23/06/1994", "Nam", "21/1 Lý Tự Trọng, P. Tân An, TP. Buôn Ma Thuột", "DN4666621963533", "01/01/2024", "31/12/2026", "66232"],
        ["26034344", "Lưu Thành Đạt", "14/06/1992", "Nam", "Thôn Sơn Cường, Xã Buôn Triết, Huyện Lắk", "DK2666623568069", "04/06/2024", "03/01/2027", "66232"],
        ["25032332", "Nguyễn Thị Tốt", "15/12/1988", "Nữ", "Tân Lợi, Buôn Ma Thuột, Đắk Lắk", "GD4666621303867", "01/01/2024", "31/12/2026", "66068"],
        ["25032336", "Ksor H' Trâm", "11/05/1984", "Nữ", "EaTam, Buôn Ma Thuột, Đắk Lắk", "GD4666621303864", "01/01/2024", "31/12/2026", "66068"],
        ["24047923", "Nguyễn Thị Hương", "12/03/1990", "Nữ", "Tân Lập, Buôn Ma Thuột, Đắk Lắk", "TE1666624722228", "01/01/2024", "31/12/2026", "66068"],
        ["20161400", "Lê Phúc Thùy Linh", "18/04/2016", "Nữ", "Tân Tiến, Buôn Ma Thuột, Đắk Lắk", "TE1676624722370", "01/01/2024", "31/12/2026", "66068"],
        ["26029578", "Nguyễn Duy Bền", "20/08/1985", "Nam", "Tự An, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""],
        ["26029570", "Hà Thị Phàng", "02/10/1975", "Nữ", "Khánh Xuân, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""],
        ["26029550", "Bùi Khắc Quý", "22/11/1968", "Nam", "Ea Kar, Đắk Lắk", "DN4666621963599", "01/01/2024", "31/12/2026", "66232"],
        ["22000008", "Bùi Minh Thiên", "14/07/1983", "Nam", "Krông Pắc, Đắk Lắk", "GD4666621303899", "01/01/2024", "31/12/2026", "66068"],
        ["19011994", "Bùi Thị Lừng", "19/01/1994", "Nữ", "Thành Nhất, Buôn Ma Thuột, Đắk Lắk", "DN4666622153400", "01/01/2024", "31/12/2026", "66232"],
        ["23027807", "Bùi Thị Thận", "05/09/1955", "Nữ", "Ea H'leo, Đắk Lắk", "", "", "", ""],
        ["26030001", "Trần Văn An", "15/03/1993", "Nam", "Tân Lợi, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""],
        ["26030002", "Lê Thị Bình", "20/05/1995", "Nữ", "Tân Lập, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""],
        ["26030003", "Phạm Văn Cường", "12/08/1989", "Nam", "Tự An, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""],
        ["26030004", "Hoàng Thị Dung", "08/11/1992", "Nữ", "Tân An, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""],
        ["26030005", "Vũ Văn Giang", "25/01/1987", "Nam", "Ea Tam, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""],
        ["26030006", "Đặng Thị Hoa", "30/04/1996", "Nữ", "Khánh Xuân, Buôn Ma Thuột, Đắk Lắk", "", "", "", ""]
    ]
    for p in patients:
        ws.append(p)
    file_path = os.path.join(CATALOG_DIR, "patients.xlsx")
    wb.save(file_path)
    print(f"Saved patients catalog template to: {file_path}")

def create_drugs_catalog():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drugs_Supplies"
    ws.append(["MaDuoc", "TenDuoc", "DVTTinh", "MaKho", "TenKho", "KhoaPhong", "Nguon", "SoLuongTon", "TrangThai", "ThoiDiemChotTon"])
    drugs = [
        ["jart1", "Jardiance 10mg", "Viên", "KHO_LE_NOITRU", "Kho lẻ Nội trú", "Khoa Nội", "BH", 1500, 1, "23/08/2026 10:00:00"],
        ["MEDT32", "MEDOLEB 200mg", "Viên", "KHO_LE_NOITRU", "Kho lẻ Nội trú", "Khoa Nội", "BH", 800, 1, "23/08/2026 10:00:00"],
        ["NEMT1", "NEXIUM MUPS 40mg", "Gói", "KHO_LE_NOITRU", "Kho lẻ Nội trú", "Khoa Nội", "BH", 1200, 1, "23/08/2026 10:00:00"],
        ["BoTV65", "Bơm tiêm MPV sử dụng 1 lần 20ml", "Cái", "KHO_VTYT_NOITRU", "Kho vật tư Nội trú", "Khoa Ngoại tổng hợp", "VP", 300, 1, "23/08/2026 10:00:00"],
        ["KiTV7", "Kim tiêm MPV", "Cái", "KHO_VTYT_NOITRU", "Kho vật tư Nội trú", "Khoa Ngoại tổng hợp", "VP", 500, 1, "23/08/2026 10:00:00"],
        ["ALLER1", "ALLERMINE 4mg", "Viên", "KHO_LE_NOITRU", "Kho lẻ Nội trú", "Khoa Nội", "VP", 600, 1, "23/08/2026 10:00:00"],
        ["BARO1", "BAROLE 20 20mg", "Viên", "KHO_LE_NOITRU", "Kho lẻ Nội trú", "Khoa Nội", "VP", 1000, 1, "23/08/2026 10:00:00"],
        ["BaTV2", "Băng thun 3 móc Urgoband 10cmx4,5m", "Cuộn", "KHO_VTYT_NOITRU", "Kho vật tư Nội trú", "Khoa Ngoại tổng hợp", "VP", 250, 1, "23/08/2026 10:00:00"],
        ["GLU24", "GLUCOSE 5%, 5%/500ml", "Chai", "KHO_LE_NOITRU", "Kho lẻ Nội trú", "Khoa Cấp cứu", "BH", 400, 1, "23/08/2026 10:00:00"],
        ["NuCT18", "NƯỚC CẤT TIÊM, 10ml, Ống", "Ống", "KHO_LE_NOITRU", "Kho lẻ Nội trú", "Khoa Cấp cứu", "BH", 3000, 1, "23/08/2026 10:00:00"],
        ["QuXT2", "Que thử nước tiểu 10 thông số", "Que", "KHO_VTYT_XN", "Kho vật tư Xét nghiệm", "Khoa Xét Nghiệm", "VP", 400, 1, "23/08/2026 10:00:00"]
    ]
    for d in drugs:
        ws.append(d)
    file_path = os.path.join(CATALOG_DIR, "drugs.xlsx")
    wb.save(file_path)
    print(f"Saved drugs catalog template to: {file_path}")

def create_services_catalog():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Services"
    ws.append(["MaDichVu", "TenDichVu", "NhomDichVu", "KhoaThucHien", "TrangThai"])
    services = [
        ["SADT", "Siêu âm Doppler tim", "Siêu âm", "Khoa Chẩn đoán hình ảnh", 1],
        ["SAOB", "Siêu âm bụng tổng quát", "Siêu âm", "Khoa Chẩn đoán hình ảnh", 1],
        ["SATG", "Siêu âm tuyến giáp", "Siêu âm", "Khoa Chẩn đoán hình ảnh", 1],
        ["SAMP", "Siêu âm màng phổi", "Siêu âm", "Khoa Chẩn đoán hình ảnh", 1],
        ["XQNT", "Chụp Xquang ngực thẳng [Số hóa 1 phim]", "X-quang", "Khoa Chẩn đoán hình ảnh", 1],
        ["XQCS", "Chụp Xquang cột sống thắt lưng thẳng nghiêng", "X-quang", "Khoa Chẩn đoán hình ảnh", 1],
        ["CTSN", "Chụp CLVT sọ não không tiêm thuốc cản quang", "CT Scan", "Khoa Chẩn đoán hình ảnh", 1],
        ["BLT", "Tổng phân tích tế bào máu ngoại vi (bằng máy đếm laser)", "Xét nghiệm", "Khoa Xét Nghiệm", 1],
        ["CRP", "CRP Định lượng [Máu]", "Xét nghiệm", "Khoa Xét Nghiệm", 1],
        ["DGD", "Điện giải đồ (Na, K, Cl) [Máu]", "Xét nghiệm", "Khoa Xét Nghiệm", 1],
        ["GLUM", "Định lượng Glucose [Máu]", "Xét nghiệm", "Khoa Xét Nghiệm", 1],
        ["CREM", "Định lượng Creatinin [Máu]", "Xét nghiệm", "Khoa Xét Nghiệm", 1],
        ["UTEN", "Tổng phân tích nước tiểu (10 thông số)", "Xét nghiệm", "Khoa Xét Nghiệm", 1],
        ["NSDD", "Nội soi dạ dày tá tràng chẩn đoán", "Nội soi", "Khoa Thăm dò chức năng", 1],
        ["DNT", "Điện tâm đồ (ECG)", "Thăm dò chức năng", "Khoa Thăm dò chức năng", 1]
    ]
    for s in services:
        ws.append(s)
    file_path = os.path.join(CATALOG_DIR, "services.xlsx")
    wb.save(file_path)
    print(f"Saved services catalog template to: {file_path}")

def create_users_catalog():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["TenDangNhap", "MatKhau", "HoTen", "KhoaPhong"])
    users = [
        ["LYCTK", "123", "Lê Yến Chi", "Khoa Nội"],
        ["TRANNNH", "123", "Trần Nguyễn Ngọc Hân", "Khoa Nội"],
        ["MENNT", "123", "Nguyễn Thị Mến", "Khoa Phụ Sản"],
        ["TOABV", "123", "Bùi Viết Tỏa", "Khoa Cấp cứu"],
        ["LACLH", "123", "Lê Anh Châu", "Khoa Ngoại tổng hợp"],
        ["NHONPT", "123", "Phạm Trọng Nhơn", "Khoa Ngoại"],
        ["NHIENVT", "123", "Nguyễn Hoàng Nhi", "Khoa Dược"],
        ["TRUCPTN", "123", "Phan Thị Ngọc Trúc", "Khoa Xét Nghiệm"],
        ["SANGVD", "123", "Vũ Đăng Sang", "Khoa Phẫu thuật - GMHS"],
        ["LINHHTT", "123", "Huỳnh Thị Linh", "Khoa Khám bệnh"],
        ["NHANHTT", "123", "Nguyễn Thanh Nhân", "Khoa Dược"]
    ]
    for u in users:
        ws.append(u)
    file_path = os.path.join(CATALOG_DIR, "users.xlsx")
    wb.save(file_path)
    print(f"Saved users catalog template to: {file_path}")

def create_service_mappings_catalog():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ServiceMappings"
    ws.append(["TenKhoaPhong", "NhomDichVu"])
    svc_maps = [
        ["Khoa Ngoại tổng hợp", "Siêu âm"],
        ["Khoa Ngoại tổng hợp", "X-quang"],
        ["Khoa Ngoại tổng hợp", "Xét nghiệm"],
        ["Khoa Ngoại tổng hợp", "CT Scan"],
        ["Khoa Ngoại tổng hợp", "Nội soi"],
        ["Khoa Nội", "Siêu âm"],
        ["Khoa Nội", "X-quang"],
        ["Khoa Nội", "Xét nghiệm"],
        ["Khoa Nội", "Thăm dò chức năng"],
        ["Khoa Cấp cứu", "Siêu âm"],
        ["Khoa Cấp cứu", "X-quang"],
        ["Khoa Cấp cứu", "Xét nghiệm"],
        ["Khoa Cấp cứu", "CT Scan"],
        ["Khoa Phụ Sản", "Siêu âm"],
        ["Khoa Phụ Sản", "Xét nghiệm"],
        ["Khoa Phẫu thuật - GMHS", "Xét nghiệm"],
        ["Khoa Phẫu thuật - GMHS", "X-quang"]
    ]
    for sm in svc_maps:
        ws.append(sm)
    file_path = os.path.join(CATALOG_DIR, "service_mappings.xlsx")
    wb.save(file_path)
    print(f"Saved service mappings template to: {file_path}")

def create_pharmacy_mappings_catalog():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PharmacyMappings"
    ws.append(["TenKhoaPhong", "MaKho", "TenKho"])
    phar_maps = [
        ["Khoa Ngoại tổng hợp", "KHO_LE_NOITRU", "Kho lẻ Nội trú"],
        ["Khoa Ngoại tổng hợp", "KHO_VTYT_NOITRU", "Kho vật tư Nội trú"],
        ["Khoa Nội", "KHO_LE_NOITRU", "Kho lẻ Nội trú"],
        ["Khoa Nội", "KHO_VTYT_NOITRU", "Kho vật tư Nội trú"],
        ["Khoa Cấp cứu", "KHO_CAPCUU", "Kho Cấp cứu"],
        ["Khoa Cấp cứu", "KHO_LE_NOITRU", "Kho lẻ Nội trú"],
        ["Khoa Phụ Sản", "KHO_LE_NOITRU", "Kho lẻ Nội trú"],
        ["Khoa Phụ Sản", "KHO_VTYT_NOITRU", "Kho vật tư Nội trú"],
        ["Khoa Phẫu thuật - GMHS", "KHO_GMHS", "Kho Phẫu thuật - GMHS"],
        ["Khoa Xét Nghiệm", "KHO_VTYT_XN", "Kho vật tư Xét nghiệm"]
    ]
    for pm in phar_maps:
        ws.append(pm)
    file_path = os.path.join(CATALOG_DIR, "pharmacy_mappings.xlsx")
    wb.save(file_path)
    print(f"Saved pharmacy mappings template to: {file_path}")

if __name__ == "__main__":
    create_folders()
    create_patients_catalog()
    create_drugs_catalog()
    create_services_catalog()
    create_users_catalog()
    create_service_mappings_catalog()
    create_pharmacy_mappings_catalog()
