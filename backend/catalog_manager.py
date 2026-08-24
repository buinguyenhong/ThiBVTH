import openpyxl
import os
import random
import json
import re
import threading
import unicodedata
from datetime import date, datetime


class PatientSelectionSession:
    """Tracks patient reservations for one exam-generation request."""

    def __init__(self, manager):
        self.manager = manager
        self.patient_keys = set()
        self.committed = False
        self.closed = False

    def commit(self):
        if self.closed:
            raise RuntimeError("Phiên chọn bệnh nhân đã đóng.")
        if not self.committed:
            self.manager._commit_patient_keys(self.patient_keys)
            self.committed = True

    def close(self):
        if self.closed:
            return
        self.manager._close_patient_selection(self)
        self.closed = True


class CatalogManager:
    def __init__(self, catalog_dir, usage_store_path=None):
        self.catalog_dir = catalog_dir
        self.usage_store_path = usage_store_path or os.path.join(
            os.path.dirname(catalog_dir), "used_patients.json"
        )
        self.patients = []
        self.drugs = []
        self.services = []
        self.users = []
        self._patient_lock = threading.RLock()
        self._patient_session = threading.local()
        self._reserved_patient_keys = set()
        self._used_patient_keys = self._load_used_patient_keys()
        
        # Load all catalogs if files exist
        self.reload_all()

    @staticmethod
    def patient_key(patient):
        """Build a stable key without depending on the randomized exam BHYT card."""
        medical_code = str(patient.get("MaYTe") or "").strip().upper()
        if medical_code:
            return f"MAYTE:{medical_code}"

        insurance_number = str(
            patient.get("SoBHYT_Nguon") or patient.get("SoBHYT") or ""
        ).strip().upper()
        if insurance_number:
            return f"BHYT:{insurance_number}"

        name = " ".join(str(patient.get("TenBenhNhan") or "").split()).upper()
        birth_date = str(patient.get("NgaySinh") or "").strip()
        return f"THONGTIN:{name}|{birth_date}"

    def _load_used_patient_keys(self):
        if not os.path.exists(self.usage_store_path):
            return set()

        try:
            with open(self.usage_store_path, "r", encoding="utf-8") as file:
                payload = json.load(file)
            keys = (
                payload.get("used_patient_keys", [])
                if isinstance(payload, dict)
                else payload
            )
            return {str(key) for key in keys if str(key).strip()}
        except Exception as exc:
            raise RuntimeError(
                f"Không đọc được lịch sử bệnh nhân đã dùng: {exc}"
            ) from exc

    def _save_used_patient_keys(self, keys):
        store_dir = os.path.dirname(self.usage_store_path) or "."
        os.makedirs(store_dir, exist_ok=True)
        temp_path = self.usage_store_path + ".tmp"
        payload = {
            "version": 1,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "used_patient_keys": sorted(keys),
        }
        try:
            with open(temp_path, "w", encoding="utf-8") as file:
                json.dump(payload, file, ensure_ascii=False, indent=2)
            os.replace(temp_path, self.usage_store_path)
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    def start_patient_selection(self):
        current = getattr(self._patient_session, "current", None)
        if current and not current.closed:
            raise RuntimeError("Đã có một phiên chọn bệnh nhân trên luồng hiện tại.")

        session = PatientSelectionSession(self)
        self._patient_session.current = session
        return session

    def _commit_patient_keys(self, patient_keys):
        if not patient_keys:
            return
        with self._patient_lock:
            updated_keys = self._used_patient_keys | set(patient_keys)
            self._save_used_patient_keys(updated_keys)
            self._used_patient_keys = updated_keys

    def _close_patient_selection(self, session):
        with self._patient_lock:
            self._reserved_patient_keys.difference_update(session.patient_keys)
        if getattr(self._patient_session, "current", None) is session:
            self._patient_session.current = None

    def get_patient_usage_status(self):
        with self._patient_lock:
            catalog_keys = {self.patient_key(patient) for patient in self.patients}
            used_in_catalog = catalog_keys & self._used_patient_keys
            return {
                "used_count": len(used_in_catalog),
                "available_count": len(catalog_keys - self._used_patient_keys),
                "used_total": len(self._used_patient_keys),
            }

    def reload_all(self):
        self.patients = self._load_patients()
        self.drugs = self._load_drugs()
        self.services = self._load_services()
        self.users = self._load_users()
        print(f"Catalogs reloaded: {len(self.patients)} patients, {len(self.drugs)} drugs, {len(self.services)} services, {len(self.users)} users")

    def _get_file_path(self, filename):
        return os.path.join(self.catalog_dir, filename)

    def _load_patients(self):
        file_path = self._get_file_path("patients.xlsx")
        if not os.path.exists(file_path):
            print(f"Warning: {file_path} not found.")
            return []
        
        patients = []
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = iter(ws.iter_rows(values_only=True))
            headers = next(rows) # Skip header
            
            for row in rows:
                if not row or not row[0]: # Skip empty rows
                    continue
                # Map row data to dict
                p = {
                    "MaYTe": str(row[0]).strip(),
                    "TenBenhNhan": str(row[1]).strip() if row[1] else "",
                    "NgaySinh": row[2] if isinstance(row[2], datetime) else (str(row[2]).strip() if row[2] else ""),
                    "GioiTinh": str(row[3]).strip() if row[3] else "Nam",
                    "DiaChiLienHe": str(row[4]).strip() if row[4] else "",
                    "SoBHYT": str(row[5]).strip() if row[5] else "",
                    "BHYTTuNgay": row[6] if isinstance(row[6], datetime) else (str(row[6]).strip() if row[6] else ""),
                    "BHYTDenNgay": row[7] if isinstance(row[7], datetime) else (str(row[7]).strip() if row[7] else ""),
                    "DKKCB": str(row[8]).strip() if (len(row) > 8 and row[8]) else "66232"
                }
                
                # Format datetime to string for easy display
                for date_key in ["NgaySinh", "BHYTTuNgay", "BHYTDenNgay"]:
                    if isinstance(p[date_key], datetime):
                        p[date_key] = p[date_key].strftime("%d/%m/%Y")
                        
                patients.append(p)
        except Exception as e:
            print(f"Error loading patients: {e}")
        finally:
            if wb is not None:
                wb.close()
        return patients

    def _load_drugs(self):
        file_path = self._get_file_path("drugs.xlsx")
        if not os.path.exists(file_path):
            print(f"Warning: {file_path} not found.")
            return []
        
        drugs = []
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = iter(ws.iter_rows(values_only=True))
            raw_headers = next(rows)
            headers = {
                str(header).strip(): index
                for index, header in enumerate(raw_headers)
                if header is not None and str(header).strip()
            }
            required_headers = {
                "MaDuoc",
                "TenDuoc",
                "DVTTinh",
                "MaKho",
                "Nguon",
                "SoLuongTon",
                "TrangThai",
            }
            missing_headers = required_headers - set(headers)
            if missing_headers:
                raise ValueError(
                    "Thiếu cột bắt buộc trong drugs.xlsx: "
                    + ", ".join(sorted(missing_headers))
                )

            def get_value(row, name, default=None):
                index = headers.get(name)
                if index is None or index >= len(row):
                    return default
                value = row[index]
                return default if value is None else value
            
            skipped_rows = []
            for row_number, row in enumerate(rows, start=2):
                if not row:
                    continue
                drug_code = get_value(row, "MaDuoc", "")
                if not drug_code:
                    continue
                try:
                    # TrangThai: 1 = Active, 0 = Inactive
                    status = int(float(get_value(row, "TrangThai", 1)))
                    if status == 0:
                        continue

                    snapshot_time = get_value(row, "ThoiDiemChotTon", "")
                    if isinstance(snapshot_time, datetime):
                        snapshot_time = snapshot_time.strftime(
                            "%d/%m/%Y %H:%M:%S"
                        )
                    else:
                        snapshot_time = (
                            str(snapshot_time).strip()
                            if snapshot_time
                            else ""
                        )

                    d = {
                        "MaDuoc": str(drug_code).strip(),
                        "TenDuoc": str(
                            get_value(row, "TenDuoc", "")
                        ).strip(),
                        "DVTTinh": str(
                            get_value(row, "DVTTinh", "")
                        ).strip(),
                        "MaKho": str(
                            get_value(row, "MaKho", "")
                        ).strip(),
                        "TenKho": str(
                            get_value(
                                row,
                                "TenKho",
                                get_value(row, "MaKho", ""),
                            )
                        ).strip(),
                        "KhoaPhong": str(
                            get_value(row, "KhoaPhong", "")
                        ).strip(),
                        "Nguon": str(
                            get_value(row, "Nguon", "VP")
                        ).strip().upper(),
                        "SoLuongTon": float(
                            get_value(row, "SoLuongTon", 0.0)
                        ),
                        "TrangThai": status,
                        "ThoiDiemChotTon": snapshot_time,
                    }
                    drugs.append(d)
                except (TypeError, ValueError) as exc:
                    skipped_rows.append((row_number, str(exc)))

            if skipped_rows:
                first_rows = ", ".join(
                    str(row_number)
                    for row_number, _ in skipped_rows[:10]
                )
                print(
                    f"Warning: skipped {len(skipped_rows)} invalid rows in "
                    f"drugs.xlsx (rows: {first_rows})."
                )
        except Exception as e:
            print(f"Error loading drugs: {e}")
        finally:
            if wb is not None:
                wb.close()
        return drugs

    def _load_services(self):
        file_path = self._get_file_path("services.xlsx")
        if not os.path.exists(file_path):
            print(f"Warning: {file_path} not found.")
            return []
        
        services = []
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = iter(ws.iter_rows(values_only=True))
            headers = next(rows)
            
            for row in rows:
                if not row or not row[0]:
                    continue
                status = int(row[4]) if row[4] is not None else 1
                if status == 0:
                    continue
                    
                s = {
                    "MaDichVu": str(row[0]).strip(),
                    "TenDichVu": str(row[1]).strip() if row[1] else "",
                    "NhomDichVu": str(row[2]).strip() if row[2] else "",
                    "KhoaThucHien": str(row[3]).strip() if row[3] else "",
                    "TrangThai": status
                }
                services.append(s)
        except Exception as e:
            print(f"Error loading services: {e}")
        finally:
            if wb is not None:
                wb.close()
        return services

    def _load_users(self):
        file_path = self._get_file_path("users.xlsx")
        if not os.path.exists(file_path):
            print(f"Warning: {file_path} not found.")
            return []
        
        users = []
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = iter(ws.iter_rows(values_only=True))
            headers = next(rows)
            
            for row in rows:
                if not row or not row[0]:
                    continue
                u = {
                    "TenDangNhap": str(row[0]).strip(),
                    "MatKhau": str(row[1]).strip() if row[1] else "",
                    "HoTen": str(row[2]).strip() if row[2] else "",
                    "KhoaPhong": str(row[3]).strip() if row[3] else ""
                }
                users.append(u)
        except Exception as e:
            print(f"Error loading users: {e}")
        finally:
            if wb is not None:
                wb.close()
        return users

    # Helper sampling methods
    @staticmethod
    def _parse_catalog_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        raw = str(value or "").strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def get_exam_insurance_period(insurance_number, exam_year):
        card = re.sub(
            r"[\s-]+", "", str(insurance_number or "").strip().upper()
        )
        if not card:
            return None, None
        if not isinstance(exam_year, int) or not 2000 <= exam_year <= 2100:
            raise ValueError("Năm thi phải nằm trong khoảng 2000 đến 2100.")

        valid_from = datetime(exam_year, 1, 1)
        valid_to = datetime(exam_year, 12, 31)
        return valid_from, valid_to

    @classmethod
    def normalize_patient_for_exam(cls, patient, exam_year):
        if not patient.get("SoBHYT"):
            return dict(patient)

        valid_from, valid_to = cls.get_exam_insurance_period(
            patient.get("SoBHYT"), exam_year
        )
        dkkcb = str(patient.get("DKKCB") or "").strip() or "66232"
        return {
            **patient,
            "BHYTTuNgay": valid_from.strftime("%d/%m/%Y"),
            "BHYTDenNgay": valid_to.strftime("%d/%m/%Y"),
            "DKKCB": dkkcb if dkkcb else "66232",
        }

    def normalize_patient_catalog(self, exam_year, file_path=None):
        """Persist exam-year BHYT periods into an uploaded patients workbook."""
        patient_file = file_path or self._get_file_path("patients.xlsx")
        workbook = None
        updated_count = 0
        try:
            workbook = openpyxl.load_workbook(patient_file)
            sheet = workbook.active
            headers = {
                str(cell.value or "").strip(): index
                for index, cell in enumerate(sheet[1], start=1)
            }
            required_headers = {"SoBHYT", "BHYTTuNgay", "BHYTDenNgay"}
            missing_headers = required_headers - set(headers)
            if missing_headers:
                raise ValueError(
                    "patients.xlsx thiếu cột: "
                    + ", ".join(sorted(missing_headers))
                )

            card_column = headers["SoBHYT"]
            valid_from_column = headers["BHYTTuNgay"]
            valid_to_column = headers["BHYTDenNgay"]
            for row_index in range(2, sheet.max_row + 1):
                insurance_number = sheet.cell(
                    row=row_index, column=card_column
                ).value
                if not str(insurance_number or "").strip():
                    continue

                valid_from, valid_to = self.get_exam_insurance_period(
                    insurance_number, exam_year
                )
                from_cell = sheet.cell(
                    row=row_index, column=valid_from_column
                )
                to_cell = sheet.cell(row=row_index, column=valid_to_column)
                from_cell.value = valid_from
                to_cell.value = valid_to
                from_cell.number_format = "dd/mm/yyyy"
                to_cell.number_format = "dd/mm/yyyy"
                updated_count += 1

            workbook.save(patient_file)
            return updated_count
        finally:
            if workbook is not None:
                workbook.close()

    def get_patients(self, count=1, must_have_bhyt=None, valid_on=None):
        """
        Retrieves a set of patients.
        If must_have_bhyt is True: returns patients with a BHYT card.
        If must_have_bhyt is False: returns patients without a BHYT card (VP).
        If valid_on is provided, BHYT cards must be valid on that date.
        """
        session = getattr(self._patient_session, "current", None)
        with self._patient_lock:
            pool = self.patients
            if must_have_bhyt is True:
                pool = [p for p in pool if p["SoBHYT"]]
            elif must_have_bhyt is False:
                pool = [p for p in pool if not p["SoBHYT"]]

            if valid_on is not None and must_have_bhyt is True:
                reference_date = self._parse_catalog_date(valid_on)
                if reference_date is None:
                    raise ValueError(
                        f"Ngày thi '{valid_on}' không đúng định dạng dd/mm/yyyy."
                    )
                pool = [
                    self.normalize_patient_for_exam(
                        patient, reference_date.year
                    )
                    for patient in pool
                ]

            if session and not session.closed:
                excluded_keys = (
                    self._used_patient_keys | self._reserved_patient_keys
                )
                unique_pool = {}
                for patient in pool:
                    key = self.patient_key(patient)
                    if key not in excluded_keys and key not in unique_pool:
                        unique_pool[key] = patient
                
                if len(unique_pool) < count:
                    for patient in pool:
                        key = self.patient_key(patient)
                        if key not in self._reserved_patient_keys and key not in unique_pool:
                            unique_pool[key] = patient

                pool = list(unique_pool.values())

            if not pool:
                criteria = (
                    "BHYT"
                    if must_have_bhyt is True
                    else "viện phí"
                    if must_have_bhyt is False
                    else ""
                )
                suffix = (
                    " chưa từng dùng"
                    if session and not session.closed
                    else ""
                )
                raise ValueError(
                    f"Không có bệnh nhân {criteria}{suffix} phù hợp trong patients.xlsx."
                )

            if session and not session.closed:
                if len(pool) < count:
                    raise ValueError(
                        "Không đủ bệnh nhân chưa từng dùng để tạo đề: "
                        f"cần {count}, còn {len(pool)} bệnh nhân phù hợp. "
                        "Hãy bổ sung patients.xlsx bằng script Lấy script."
                    )
                selected = random.sample(pool, k=count)
                selected_keys = {
                    self.patient_key(patient) for patient in selected
                }
                session.patient_keys.update(selected_keys)
                self._reserved_patient_keys.update(selected_keys)
                return selected

            if len(pool) < count:
                # Calls outside an exam transaction keep legacy sampling behavior.
                return random.choices(pool, k=count)
            return random.sample(pool, k=count)

    def get_service_groups(self):
        """Returns a sorted list of unique service groups (NhomDichVu) from catalog."""
        groups = set()
        for s in self.services:
            g = str(s.get("NhomDichVu", "")).strip()
            if g:
                groups.add(g)
        return sorted(list(groups))

    def get_warehouses(self):
        """Returns a sorted list of unique warehouse codes (MaKho) from catalog."""
        whs = set()
        for d in self.drugs:
            w = str(d.get("MaKho", "")).strip()
            if w:
                whs.add(w)
        return sorted(list(whs))

    def get_departments_from_users(self):
        """Returns a sorted list of unique departments from users catalog."""
        depts = set()
        for u in self.users:
            dp = str(u.get("KhoaPhong", "")).strip()
            if dp:
                depts.add(dp)
        return sorted(list(depts))

    def get_drugs(self, count=1, nguon=None, min_ton=1.0, exclude_codes=None, khoa=None, warehouses=None):
        """
        Retrieves drugs from the catalog matching source, stock, department, and warehouse criteria.
        """
        pool = self.drugs
        if nguon:
            pool = [d for d in pool if d["Nguon"].upper() == nguon.upper()]
        if min_ton:
            pool = [d for d in pool if d["SoLuongTon"] >= min_ton]
        if exclude_codes:
            pool = [d for d in pool if d["MaDuoc"] not in exclude_codes]
        if warehouses and isinstance(warehouses, list) and len(warehouses) > 0:
            wh_filtered = [d for d in pool if d.get("MaKho") in warehouses]
            if wh_filtered:
                pool = wh_filtered
        elif khoa:
            requested_department = self._normalize_department(khoa)
            khoa_filtered = [
                drug
                for drug in pool
                if drug.get("KhoaPhong")
                and self._normalize_department(drug["KhoaPhong"])
                == requested_department
            ]
            if not khoa_filtered:
                normalized_keyword = self._normalize_text(khoa)
                khoa_filtered = [
                    drug
                    for drug in pool
                    if normalized_keyword
                    and normalized_keyword
                    in self._normalize_text(drug.get("TenDuoc"))
                ]
            if khoa_filtered:
                pool = khoa_filtered
            
        if not pool:
            pool = self.drugs
            
        if len(pool) < count:
            return random.choices(pool, k=count) if pool else []
        return random.sample(pool, k=count)

    def get_services(self, count=1, nhom=None, khoa=None):
        """
        Retrieves clinical services matching category and executing department.
        """
        pool = self.services
        if nhom:
            if isinstance(nhom, list):
                pool = [s for s in pool if s["NhomDichVu"] in nhom]
            else:
                pool = [s for s in pool if s["NhomDichVu"] == nhom]
        if khoa:
            khoa_lower = khoa.lower()
            khoa_filtered = [s for s in pool if khoa_lower in s["KhoaThucHien"].lower() or khoa_lower in s["NhomDichVu"].lower()]
            if khoa_filtered:
                pool = khoa_filtered
            
        if not pool:
            pool = self.services
            
        if len(pool) < count:
            return random.choices(pool, k=count) if pool else []
        return random.sample(pool, k=count)

    def get_user_for_dept(self, dept_name):
        """
        Finds a user account that corresponds to the given department name.
        """
        pool = self.get_users_for_dept(dept_name)
        if not pool:
            raise ValueError(
                f"Không có user đang hoạt động thuộc đúng khoa/phòng "
                f"'{dept_name}' trong users.xlsx. Hãy dùng nút Lấy script "
                "để cập nhật danh mục user theo khoa/phòng."
            )
        return random.choice(pool)

    def get_users_for_dept(self, dept_name):
        """Returns all active catalog users assigned to a department."""
        requested_department = self._normalize_department(dept_name)
        return [
            user
            for user in self.users
            if self._normalize_department(user.get("KhoaPhong"))
            == requested_department
        ]

    def has_user_for_dept(self, dept_name):
        """Checks whether the user catalog can serve a department."""
        return bool(self.get_users_for_dept(dept_name))

    @staticmethod
    def _normalize_text(value):
        raw = " ".join(str(value or "").strip().casefold().split())
        decomposed = unicodedata.normalize("NFD", raw)
        return re.sub(
            r"[^a-z0-9]+",
            " ",
            "".join(
                char
                for char in decomposed
                if unicodedata.category(char) != "Mn"
            ).replace("đ", "d"),
        ).strip()

    @classmethod
    def _normalize_department(cls, value):
        text = cls._normalize_text(value)

        for prefix in ("khoa ", "phong ", "quay "):
            if text.startswith(prefix):
                text = text[len(prefix):].strip()

        aliases = {
            "cdha": "chan doan hinh anh",
            "hstc": "hoi suc tich cuc",
            "tmh": "tai mui hong",
            "rhm": "rang ham mat",
            "xn": "xet nghiem",
            "ngoai th": "ngoai tong hop",
            "ngoai ctch": "ngoai chan thuong chinh hinh",
            "ngoai tk": "ngoai than kinh",
        }
        return aliases.get(text, text)

    def get_inventory_options(self):
        """Returns department/warehouse filters present in the uploaded catalog."""
        departments = sorted(
            {
                drug["KhoaPhong"]
                for drug in self.drugs
                if drug.get("KhoaPhong")
            },
            key=self._normalize_text,
        )
        warehouse_map = {}
        for drug in self.drugs:
            warehouse_code = drug.get("MaKho", "")
            if not warehouse_code:
                continue
            warehouse_map.setdefault(
                warehouse_code,
                {
                    "code": warehouse_code,
                    "name": drug.get("TenKho") or warehouse_code,
                    "department": drug.get("KhoaPhong") or "",
                },
            )

        snapshot_times = sorted(
            {
                drug["ThoiDiemChotTon"]
                for drug in self.drugs
                if drug.get("ThoiDiemChotTon")
            }
        )
        return {
            "departments": departments,
            "warehouses": sorted(
                warehouse_map.values(),
                key=lambda item: (
                    self._normalize_text(item["department"]),
                    self._normalize_text(item["name"]),
                    item["code"],
                ),
            ),
            "sources": sorted(
                {
                    drug["Nguon"]
                    for drug in self.drugs
                    if drug.get("Nguon")
                }
            ),
            "snapshot_times": snapshot_times,
            "has_department_mapping": bool(departments),
        }

    def search_inventory(
        self,
        department=None,
        warehouse=None,
        query=None,
        source=None,
        min_stock=0.0,
        limit=200,
        mapped_warehouses=None,
    ):
        """Filters the uploaded stock snapshot without querying HIS."""
        pool = list(self.drugs)

        if department:
            normalized_department = self._normalize_department(department)
            wh_set = {str(w).strip().casefold() for w in (mapped_warehouses or []) if w}
            pool = [
                drug
                for drug in pool
                if (
                    (drug.get("KhoaPhong") and self._normalize_department(drug.get("KhoaPhong")) == normalized_department)
                    or (str(drug.get("MaKho") or "").strip().casefold() in wh_set)
                    or (str(drug.get("TenKho") or "").strip().casefold() in wh_set)
                )
            ]
        if warehouse:
            normalized_warehouse = str(warehouse).strip().casefold()
            pool = [
                drug
                for drug in pool
                if str(drug.get("MaKho") or "").strip().casefold() == normalized_warehouse
                or str(drug.get("TenKho") or "").strip().casefold() == normalized_warehouse
            ]
        if source:
            normalized_source = str(source).strip().upper()
            pool = [
                drug
                for drug in pool
                if str(drug.get("Nguon") or "").strip().upper() == normalized_source
            ]

        pool = [
            drug
            for drug in pool
            if float(drug.get("SoLuongTon") or 0) >= float(min_stock)
        ]

        normalized_query = self._normalize_text(query)
        if normalized_query:
            searchable_fields = (
                "MaDuoc",
                "TenDuoc",
                "MaKho",
                "TenKho",
                "KhoaPhong",
            )
            pool = [
                drug
                for drug in pool
                if any(
                    normalized_query
                    in self._normalize_text(drug.get(field))
                    for field in searchable_fields
                )
            ]

        pool.sort(
            key=lambda drug: (
                self._normalize_text(drug.get("KhoaPhong")),
                self._normalize_text(drug.get("TenKho")),
                self._normalize_text(drug.get("TenDuoc")),
                drug.get("Nguon", ""),
            )
        )
        total = len(pool)
        returned_items = pool[:limit]
        return {
            "items": returned_items,
            "total": total,
            "returned": len(returned_items),
            "total_stock": sum(
                float(drug.get("SoLuongTon") or 0) for drug in pool
            ),
        }

# Test block
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    cm = CatalogManager(r"C:\Users\Admin\.gemini\antigravity\scratch\ExamGenerator\data\catalogs")
    print("Testing get_patients:")
    print(cm.get_patients(2, must_have_bhyt=True))
    print("\nTesting get_drugs:")
    print(cm.get_drugs(2, nguon="BH"))
    print("\nTesting get_services:")
    print(cm.get_services(2, nhom="Siêu âm"))
    print("\nTesting get_user_for_dept:")
    print(cm.get_user_for_dept("Nội"))
