import json
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BACKEND_DIR))

from catalog_manager import CatalogManager


class CatalogManagerPatientUsageTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        data_dir = Path(self.temp_dir.name)
        catalog_dir = data_dir / "catalogs"
        catalog_dir.mkdir()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "MaYTe",
                "TenBenhNhan",
                "NgaySinh",
                "GioiTinh",
                "DiaChiLienHe",
                "SoBHYT",
                "BHYTTuNgay",
                "BHYTDenNgay",
                "DKKCB",
            ]
        )
        sheet.append(
            [
                "BH001",
                "Bệnh nhân BHYT 1",
                "1990-01-01",
                "Nam",
                "Đắk Lắk",
                "DN4666620000001",
                "2024-01-01",
                "2030-12-31",
                "66232",
            ]
        )
        sheet.append(
            [
                "BH002",
                "Bệnh nhân BHYT 2",
                "1991-01-01",
                "Nữ",
                "Đắk Lắk",
                "TE1666620000002",
                "2020-01-01",
                "2020-12-31",
                "66232",
            ]
        )
        sheet.append(
            [
                "VP001",
                "Bệnh nhân viện phí 1",
                "1992-01-01",
                "Nam",
                "Đắk Lắk",
                "",
                "",
                "",
                "",
            ]
        )
        sheet.append(
            [
                "VP002",
                "Bệnh nhân viện phí 2",
                "1993-01-01",
                "Nữ",
                "Đắk Lắk",
                "",
                "",
                "",
                "",
            ]
        )
        workbook.save(catalog_dir / "patients.xlsx")

        user_workbook = openpyxl.Workbook()
        user_sheet = user_workbook.active
        user_sheet.append(["TenDangNhap", "MatKhau", "HoTen", "KhoaPhong"])
        user_sheet.append(["USER_NOI", "123", "Người dùng Nội", "Khoa Nội"])
        user_sheet.append(["USER_CDHA", "123", "Người dùng CĐHA", "CĐHA"])
        user_workbook.save(catalog_dir / "users.xlsx")

        drug_workbook = openpyxl.Workbook()
        drug_sheet = drug_workbook.active
        drug_sheet.append(
            [
                "MaDuoc",
                "TenDuoc",
                "DVTTinh",
                "MaKho",
                "TenKho",
                "KhoaPhong",
                "Nguon",
                "SoLuongTon",
                "TrangThai",
                "ThoiDiemChotTon",
            ]
        )
        drug_sheet.append(
            [
                "THUOC01",
                "Amoxicillin 500mg",
                "Viên",
                "KNOI",
                "Kho Khoa Nội",
                "Khoa Nội",
                "BH",
                125.5,
                1,
                "28/07/2026 08:30:00",
            ]
        )
        drug_sheet.append(
            [
                "VTYT01",
                "Ống nghiệm xét nghiệm",
                "Cái",
                "KXN",
                "Kho Xét nghiệm",
                "Khoa Xét Nghiệm",
                "VP",
                45,
                1,
                "28/07/2026 08:30:00",
            ]
        )
        drug_workbook.save(catalog_dir / "drugs.xlsx")

        self.usage_path = data_dir / "used_patients.json"
        self.catalog_dir = catalog_dir
        self.manager = CatalogManager(
            str(catalog_dir), usage_store_path=str(self.usage_path)
        )

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_committed_patients_are_not_selected_again(self):
        first_session = self.manager.start_patient_selection()
        try:
            first_bhyt = self.manager.get_patients(
                1, must_have_bhyt=True, valid_on="27/07/2026"
            )[0]
            first_vp = self.manager.get_patients(1, must_have_bhyt=False)[0]
            first_session.commit()
        finally:
            first_session.close()

        second_session = self.manager.start_patient_selection()
        try:
            second_bhyt = self.manager.get_patients(
                1, must_have_bhyt=True, valid_on="27/07/2026"
            )[0]
            second_vp = self.manager.get_patients(1, must_have_bhyt=False)[0]
            second_session.commit()
        finally:
            second_session.close()

        self.assertNotEqual(first_bhyt["MaYTe"], second_bhyt["MaYTe"])
        self.assertNotEqual(first_vp["MaYTe"], second_vp["MaYTe"])
        payload = json.loads(self.usage_path.read_text(encoding="utf-8"))
        self.assertEqual(len(payload["used_patient_keys"]), 4)
        self.assertEqual(self.manager.get_patient_usage_status()["available_count"], 0)

    def test_failed_session_releases_patient_without_marking_it_used(self):
        failed_session = self.manager.start_patient_selection()
        try:
            first = self.manager.get_patients(1, must_have_bhyt=False)[0]
        finally:
            failed_session.close()

        retry_session = self.manager.start_patient_selection()
        try:
            available_codes = {
                patient["MaYTe"]
                for patient in self.manager.get_patients(
                    2, must_have_bhyt=False
                )
            }
        finally:
            retry_session.close()

        self.assertIn(first["MaYTe"], available_codes)
        self.assertFalse(self.usage_path.exists())

    def test_user_must_match_requested_department(self):
        internal_medicine_user = self.manager.get_user_for_dept("Nội")
        imaging_user = self.manager.get_user_for_dept(
            "Khoa Chẩn đoán hình ảnh"
        )

        self.assertEqual(internal_medicine_user["TenDangNhap"], "USER_NOI")
        self.assertEqual(imaging_user["TenDangNhap"], "USER_CDHA")
        self.assertTrue(self.manager.has_user_for_dept("Khoa Nội"))
        self.assertFalse(self.manager.has_user_for_dept("Khoa Ngoại"))
        with self.assertRaisesRegex(ValueError, "đúng khoa/phòng"):
            self.manager.get_user_for_dept("Khoa Ngoại")

    def test_insurance_period_is_normalized_for_exam_year(self):
        updated_count = self.manager.normalize_patient_catalog(2026)
        self.manager.reload_all()

        patients_by_code = {
            patient["MaYTe"]: patient for patient in self.manager.patients
        }
        standard_card = patients_by_code["BH001"]
        child_card = patients_by_code["BH002"]

        self.assertEqual(updated_count, 2)
        self.assertEqual(standard_card["BHYTTuNgay"], "01/01/2026")
        self.assertEqual(standard_card["BHYTDenNgay"], "31/12/2026")
        self.assertEqual(child_card["BHYTTuNgay"], "01/01/2026")
        self.assertEqual(child_card["BHYTDenNgay"], "31/12/2026")

        selected = self.manager.get_patients(
            2, must_have_bhyt=True, valid_on="15/07/2026"
        )
        selected_by_code = {
            patient["MaYTe"]: patient for patient in selected
        }
        self.assertEqual(
            selected_by_code["BH001"]["BHYTDenNgay"], "31/12/2026"
        )
        self.assertEqual(
            selected_by_code["BH002"]["BHYTDenNgay"], "31/12/2026"
        )

    def test_inventory_can_be_filtered_by_department_and_text(self):
        options = self.manager.get_inventory_options()
        self.assertTrue(options["has_department_mapping"])
        self.assertEqual(
            options["snapshot_times"], ["28/07/2026 08:30:00"]
        )
        self.assertIn("Khoa Nội", options["departments"])
        self.assertIn("Khoa Xét Nghiệm", options["departments"])

        result = self.manager.search_inventory(
            department="Xét nghiệm",
            query="ong nghiem",
            min_stock=10,
        )
        self.assertEqual(result["total"], 1)
        self.assertEqual(result["items"][0]["MaDuoc"], "VTYT01")
        self.assertEqual(result["total_stock"], 45)

    def test_legacy_seven_column_drug_catalog_remains_compatible(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "MaDuoc",
                "TenDuoc",
                "DVTTinh",
                "MaKho",
                "Nguon",
                "SoLuongTon",
                "TrangThai",
            ]
        )
        sheet.append(["OLD01", "Dược cũ", "Viên", "KHO01", "BH", 12, 1])
        workbook.save(self.catalog_dir / "drugs.xlsx")

        self.manager.reload_all()

        self.assertEqual(len(self.manager.drugs), 1)
        self.assertEqual(self.manager.drugs[0]["TenKho"], "KHO01")
        self.assertEqual(self.manager.drugs[0]["KhoaPhong"], "")
        self.assertFalse(
            self.manager.get_inventory_options()["has_department_mapping"]
        )


if __name__ == "__main__":
    unittest.main()
