import os
import json
import threading

class MappingManager:
    """Manages Department-to-ServiceGroup mappings and Department-to-Pharmacy mappings."""

    def __init__(self, config_dir=None):
        if config_dir is None:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_dir = os.path.join(base_dir, "data", "config")
        self.config_dir = config_dir
        os.makedirs(self.config_dir, exist_ok=True)
        
        self.service_mapping_path = os.path.join(self.config_dir, "service_mappings.json")
        self.pharmacy_mapping_path = os.path.join(self.config_dir, "pharmacy_mappings.json")
        self.warehouse_names_path = os.path.join(self.config_dir, "warehouse_names.json")
        self._lock = threading.RLock()
        
        self.service_mappings = self._load_json(self.service_mapping_path, default=self._default_service_mappings())
        self.pharmacy_mappings = self._load_json(self.pharmacy_mapping_path, default=self._default_pharmacy_mappings())
        self.warehouse_names = self._load_json(self.warehouse_names_path, default=self._default_warehouse_names())

    def _load_json(self, path, default):
        if not os.path.exists(path):
            self._save_json(path, default)
            return dict(default)
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    # Ensure all standard departments exist by merging defaults
                    merged = dict(default)
                    merged.update(data)
                    return merged
                return dict(default)
        except Exception as e:
            print(f"Error loading {path}: {e}, using default.")
            return dict(default)

    def _save_json(self, path, data):
        temp_path = path + ".tmp"
        try:
            with open(temp_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            os.replace(temp_path, path)
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    def _default_service_mappings(self):
        # Default mappings for known hospital departments
        return {
            "Khoa Ngoại tổng hợp": ["Siêu âm", "X-quang", "CT Scan", "Xét nghiệm", "Thủ thuật"],
            "Khoa Chấn thương chỉnh hình": ["X-quang", "CT Scan", "Xét nghiệm", "Siêu âm", "Thủ thuật"],
            "Khoa Ngoại thần kinh": ["CT Scan", "MRI", "X-quang", "Xét nghiệm", "Siêu âm"],
            "Khoa Hồi sức tích cực": ["Xét nghiệm", "X-quang", "Siêu âm", "Khí máu"],
            "Khoa Nội": ["Xét nghiệm", "Siêu âm", "X-quang", "Nội soi", "Điện tim"],
            "Khoa Tim mạch": ["Điện tim", "Siêu âm tim", "Xét nghiệm", "X-quang", "Holter"],
            "Khoa Phụ Sản": ["Siêu âm", "Xét nghiệm", "Monitor sản khoa", "Thủ thuật"],
            "Khoa Nhi": ["Xét nghiệm", "X-quang", "Siêu âm"],
            "Khoa Cấp cứu": ["Xét nghiệm", "X-quang", "Siêu âm", "CT Scan", "Điện tim"],
            "Khoa Mắt": ["Khám mắt", "Đo khúc xạ", "Thủ thuật mắt", "Xét nghiệm"],
            "Khoa Tai Mũi Họng": ["Nội soi TMH", "Thủ thuật TMH", "Xét nghiệm", "X-quang"],
            "Khoa Răng Hàm Mặt": ["X-quang RHM", "Thủ thuật RHM", "Xét nghiệm"],
            "Khoa Khám bệnh": ["Khám bệnh", "Xét nghiệm", "Siêu âm", "X-quang", "Điện tim"],
            "Thận nhân tạo": ["Xét nghiệm", "Lọc máu", "Điện tim", "Thủ thuật"],
            "Khoa Phục hồi chức năng": ["Vật lý trị liệu", "Phục hồi chức năng", "Thủ thuật", "Điện xung", "Siêu âm điều trị", "Hồng ngoại", "Laser điều trị"],
            "Khoa Xét Nghiệm": ["Xét nghiệm", "Huyết học", "Hóa sinh", "Vi sinh"],
            "Khoa Chẩn đoán hình ảnh": ["X-quang", "Siêu âm", "CT Scan", "MRI"],
            "Khoa Nội soi": ["Nội soi tiêu hóa", "Nội soi dạ dày", "Nội soi đại tràng"],
            "Khoa Phẫu thuật - GMHS": ["Xét nghiệm", "Thủ thuật", "Gây mê hồi sức"],
            "Khoa Dược": ["Dược lâm sàng", "Kiểm nghiệm"],
            "Phòng Tài chính kế toán": [],
            "Phòng Chăm sóc khách hàng": [],
            "Phòng Tổ chức - Hành chính": []
        }

    def _default_pharmacy_mappings(self):
        # Default mapping of department to warehouse codes
        return {
            "Khoa Ngoại tổng hợp": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_NGOAI", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Khoa Chấn thương chỉnh hình": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_CTCH", "KHO_NGOAI", "KHO_CHUNG"],
            "Khoa Ngoại thần kinh": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_NGOAITK", "KHO_NGOAI", "KHO_CHUNG"],
            "Khoa Hồi sức tích cực": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_HSTC", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Khoa Nội": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_NOI", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Khoa Tim mạch": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_TIMMACH", "KHO_NOI", "KHO_CHUNG"],
            "Khoa Phụ Sản": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_SAN", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Khoa Nhi": ["KHO_LE_NOITRU", "KHO_VTYT_NOITRU", "KHO_NHI", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Khoa Cấp cứu": ["KHO_LE", "KHO_CAPCUU", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Khoa Mắt": ["KHO_LE", "KHO_MAT", "KHO_PK", "KHO_CHUNG"],
            "Khoa Tai Mũi Họng": ["KHO_LE", "KHO_TMH", "KHO_PK", "KHO_CHUNG"],
            "Khoa Răng Hàm Mặt": ["KHO_LE", "KHO_RHM", "KHO_PK", "KHO_CHUNG"],
            "Khoa Khám bệnh": ["KHO_LE", "KHO_BHYT", "KHO_PK", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Thận nhân tạo": ["KHO_LE_NOITRU", "KHO_TNT", "KHO_CHUNG"],
            "Khoa Phục hồi chức năng": ["KHO_LE", "KHO_PK", "KHO_CHUNG", "KHO_DUOCTONG"],
            "Khoa Phẫu thuật - GMHS": ["KHO_PTGM", "KHO_GMHS", "KHO_CHUNG"],
            "Khoa Dược": ["KHO_DUOCTONG", "KHO_CHUNG", "KHO_BHYT", "KHO_LE"]
        }

    def _default_warehouse_names(self):
        return {
            "KHO_LE": "Kho lẻ Dược",
            "KHO_BHYT": "Kho BHYT",
            "KHO_LE_NOITRU": "Kho lẻ Nội trú",
            "KHO_VTYT_NOITRU": "Kho vật tư Nội trú",
            "KHO_NGOAI": "Kho Dược Ngoại",
            "KHO_NOI": "Kho Dược Nội",
            "KHO_CHUNG": "Kho Dược dùng chung",
            "KHO_DUOCTONG": "Kho Dược Tổng",
            "KHO_CAPCUU": "Kho Cấp cứu",
            "KHO_HSTC": "Kho Hồi sức tích cực",
            "KHO_CTCH": "Kho Chấn thương chỉnh hình",
            "KHO_NGOAITK": "Kho Ngoại thần kinh",
            "KHO_TIMMACH": "Kho Tim mạch",
            "KHO_SAN": "Kho Phụ Sản",
            "KHO_NHI": "Kho Nhi",
            "KHO_MAT": "Kho Mắt",
            "KHO_TMH": "Kho Tai Mũi Họng",
            "KHO_RHM": "Kho Răng Hàm Mặt",
            "KHO_PK": "Kho Phòng khám",
            "KHO_TNT": "Kho Thận nhân tạo",
            "KHO_PTGM": "Kho Phẫu thuật - GMHS",
            "KHO_GMHS": "Kho Gây mê hồi sức",
            "KHO_VTYT_XN": "Kho VTYT Xét nghiệm"
        }

    def get_warehouse_name(self, code):
        if not code:
            return ""
        with self._lock:
            return self.warehouse_names.get(code, code)

    def get_all_warehouse_metadata(self, catalog_warehouses=None):
        """Returns list of warehouse dicts: [{code, name, display}, ...]"""
        with self._lock:
            all_codes = set(self.warehouse_names.keys())
            if catalog_warehouses:
                all_codes.update(catalog_warehouses)
            for wh_list in self.pharmacy_mappings.values():
                all_codes.update(wh_list)
            
            result = []
            for code in sorted(list(all_codes)):
                name = self.warehouse_names.get(code, code)
                display = f"{name} ({code})" if name != code else code
                result.append({
                    "code": code,
                    "name": name,
                    "display": display
                })
            return result

    def get_services_for_dept(self, dept_name):
        with self._lock:
            return self.service_mappings.get(dept_name, [])

    def set_services_for_dept(self, dept_name, service_groups):
        with self._lock:
            self.service_mappings[dept_name] = sorted(list(set(service_groups)))
            self._save_json(self.service_mapping_path, self.service_mappings)

    def get_all_service_mappings(self):
        with self._lock:
            return dict(self.service_mappings)

    def update_all_service_mappings(self, mappings):
        with self._lock:
            self.service_mappings = mappings
            self._save_json(self.service_mapping_path, self.service_mappings)

    def get_pharmacies_for_dept(self, dept_name):
        with self._lock:
            return self.pharmacy_mappings.get(dept_name, [])

    def set_pharmacies_for_dept(self, dept_name, warehouse_codes):
        with self._lock:
            self.pharmacy_mappings[dept_name] = sorted(list(set(warehouse_codes)))
            self._save_json(self.pharmacy_mapping_path, self.pharmacy_mappings)

    def get_all_pharmacy_mappings(self):
        with self._lock:
            return dict(self.pharmacy_mappings)

    def update_all_pharmacy_mappings(self, mappings):
        with self._lock:
            self.pharmacy_mappings = mappings
            self._save_json(self.pharmacy_mapping_path, self.pharmacy_mappings)

    def import_services_from_excel(self, file_path):
        """Imports department-service group mappings from an Excel file exported by Script 07."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
        dept_idx = -1
        group_idx = -1
        
        for idx, h in enumerate(headers):
            if any(k in h for k in ["tenkhoaphong", "tenphongban", "khoa", "phongban"]):
                dept_idx = idx
            elif any(k in h for k in ["nhomdichvu", "nhom", "group"]):
                group_idx = idx

        if dept_idx == -1 or group_idx == -1:
            raise ValueError("File Excel cần có các cột: TenKhoaPhong và NhomDichVu.")

        new_map = {}
        row_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(dept_idx, group_idx):
                continue
            dept = str(row[dept_idx] or "").strip()
            grp = str(row[group_idx] or "").strip()
            if not dept or not grp or grp.lower() == "none":
                continue
            if dept not in new_map:
                new_map[dept] = set()
            new_map[dept].add(grp)
            row_count += 1

        with self._lock:
            for dept, grps in new_map.items():
                self.service_mappings[dept] = sorted(list(grps))
            self._save_json(self.service_mapping_path, self.service_mappings)

        return len(new_map), row_count

    def import_pharmacies_from_excel(self, file_path):
        """Imports department-pharmacy warehouse mappings from an Excel file exported by Script 08."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
        dept_idx = -1
        wh_idx = -1
        name_idx = -1
        
        for idx, h in enumerate(headers):
            if any(k in h for k in ["tenkhoaphong", "tenphongban", "khoa", "phongban"]):
                dept_idx = idx
            elif "makho" in h or h == "kho" or "warehouse" in h or "ma_kho" in h:
                wh_idx = idx
            elif any(k in h for k in ["tenkho", "ten_kho", "name", "warehouse_name"]):
                name_idx = idx

        if dept_idx == -1 or wh_idx == -1:
            raise ValueError("File Excel cần có các cột: TenKhoaPhong và MaKho.")

        new_map = {}
        row_count = 0
        with self._lock:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(dept_idx, wh_idx):
                    continue
                dept = str(row[dept_idx] or "").strip()
                wh = str(row[wh_idx] or "").strip()
                if not dept or not wh or wh.lower() == "none":
                    continue
                if name_idx != -1 and len(row) > name_idx and row[name_idx]:
                    wh_name = str(row[name_idx]).strip()
                    if wh_name:
                        self.warehouse_names[wh] = wh_name
                if dept not in new_map:
                    new_map[dept] = set()
                new_map[dept].add(wh)
                row_count += 1

            for dept, whs in new_map.items():
                self.pharmacy_mappings[dept] = sorted(list(whs))
            self._save_json(self.pharmacy_mapping_path, self.pharmacy_mappings)
            self._save_json(self.warehouse_names_path, self.warehouse_names)

        return len(new_map), row_count
